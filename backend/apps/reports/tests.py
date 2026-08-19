from io import BytesIO
from datetime import date, datetime
from decimal import Decimal

from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.distribution.models import Distribution
from apps.items.models import Category, Facility, FundingSource, Item, Location, Supplier, Unit
from apps.procurement.models import ProcurementContract
from apps.receiving.models import Receiving, ReceivingItem
from apps.reports.exports import (
	export_numbering_history_excel,
	export_pengeluaran_excel,
	export_rekap_excel,
)
from apps.stock.models import OpeningBalanceImport, OpeningBalanceImportItem, Stock, Transaction
from apps.users.models import User


class NumberingHistoryReportTests(TestCase):
	@classmethod
	def setUpTestData(cls):
		cls.user = User.objects.create_superuser(
			username="reports-admin",
			password="secret12345",
		)
		cls.unit = Unit.objects.create(code="TAB", name="Tablet")
		cls.category = Category.objects.create(
			code="REPORT-CAT", name="Report Category", sort_order=1
		)
		cls.item = Item.objects.create(
			nama_barang="Paracetamol 500mg",
			satuan=cls.unit,
			kategori=cls.category,
		)
		cls.location = Location.objects.create(code="REP-LOC", name="Gudang Laporan")
		cls.funding_source = FundingSource.objects.create(code="BOK", name="BOK")
		cls.facility = Facility.objects.create(code="PKM-REP", name="Puskesmas Laporan")
		cls.stock = Stock.objects.create(
			item=cls.item,
			location=cls.location,
			batch_lot="REP-01",
			expiry_date="2027-12-31",
			quantity=10,
			reserved=0,
			unit_price=1000,
			sumber_dana=cls.funding_source,
		)

	def setUp(self):
		self.client.force_login(self.user)

	def _create_distribution(self, distribution_type, document_number=None):
		dist = Distribution.objects.create(
			distribution_type=distribution_type,
			document_number=document_number or "",
			request_date="2026-04-01",
			facility=self.facility,
			status=Distribution.Status.DRAFT,
			created_by=self.user,
			notes="Catatan ringkas",
		)
		dist.items.create(
			item=self.item,
			quantity_requested=5,
			quantity_approved=5,
			stock=self.stock,
		)
		return dist

	def test_numbering_history_page_lists_lplpo_and_special_request(self):
		lplpo_dist = self._create_distribution(Distribution.DistributionType.LPLPO)
		special_dist = self._create_distribution(Distribution.DistributionType.SPECIAL_REQUEST)

		response = self.client.get(reverse('reports:numbering_history'), secure=True)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, lplpo_dist.document_number)
		self.assertContains(response, special_dist.document_number)
		self.assertContains(response, "Riwayat Penomoran")
		self.assertContains(response, "Lihat Dokumen")

	def test_numbering_history_page_filters_by_document_type(self):
		lplpo_dist = self._create_distribution(Distribution.DistributionType.LPLPO)
		self._create_distribution(Distribution.DistributionType.SPECIAL_REQUEST)

		response = self.client.get(
			reverse('reports:numbering_history'),
			{'distribution_type': Distribution.DistributionType.LPLPO, 'year': 2026},
			secure=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, lplpo_dist.document_number)
		self.assertNotContains(response, 'KD.F/2026')

	def test_numbering_history_page_shows_print_and_export_actions(self):
		self._create_distribution(Distribution.DistributionType.LPLPO)

		response = self.client.get(reverse('reports:numbering_history'), secure=True)

		self.assertContains(response, 'Cetak Laporan')
		self.assertContains(response, 'Export Excel')

	def test_numbering_history_excel_export_returns_workbook(self):
		self._create_distribution(Distribution.DistributionType.LPLPO)

		response = self.client.get(
			reverse('reports:numbering_history'),
			{'year': 2026, 'format': 'excel'},
			secure=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(
			response['Content-Type'],
			'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		)
		self.assertIn('Riwayat_Penomoran_2026.xlsx', response['Content-Disposition'])

	def test_numbering_history_excel_neutralizes_formula_prefixed_strings(self):
		response = export_numbering_history_excel(
			[
				{
					"document_number": "=DOC-001",
					"distribution_type": "+LPLPO",
					"status": "@Draft",
					"facility_name": "-Facility",
					"source_label": "=LPLPO",
					"source_document_number": "=SRC-001",
					"created_at": None,
					"item_count": 1,
				}
			],
			2026,
			"=Semua Dokumen",
			)

		workbook = load_workbook(BytesIO(response.content))
		sheet = workbook.active

		self.assertEqual(sheet["A2"].value, "Tahun: 2026 | Jenis Dokumen: =Semua Dokumen")
		self.assertEqual(sheet["B5"].value, "'=DOC-001")
		self.assertEqual(sheet["C5"].value, "'+LPLPO")
		self.assertEqual(sheet["D5"].value, "'@Draft")
		self.assertEqual(sheet["E5"].value, "'-Facility")
		self.assertEqual(sheet["F5"].value, "'=LPLPO: =SRC-001")
		self.assertEqual(sheet["A2"].data_type, "s")
		self.assertEqual(sheet["B5"].data_type, "s")


@override_settings(SECURE_SSL_REDIRECT=False, ALLOWED_HOSTS=['testserver', 'localhost', '127.0.0.1'])
class RekapOpeningBalanceReportTests(TestCase):
	@classmethod
	def setUpTestData(cls):
		cls.user = User.objects.create_superuser(
			username="rekap-opening-admin",
			password="secret12345",
		)
		cls.unit = Unit.objects.create(code="RO-TAB", name="Tablet")
		cls.category = Category.objects.create(
			code="RO-CAT", name="Obat Rekap", sort_order=1
		)
		cls.item = Item.objects.create(
			kode_barang="RO-ITEM-001",
			nama_barang="Item Rekap Saldo Awal",
			satuan=cls.unit,
			kategori=cls.category,
		)
		cls.location = Location.objects.create(code="RO-LOC", name="Gudang Rekap")
		cls.funding = FundingSource.objects.create(code="RO-FUND", name="Dana Rekap")
		cls.opening_balance = OpeningBalanceImport.objects.create(
			document_number="SALDO-AWAL-2026-REKAP",
			effective_date=date(2026, 1, 1),
			created_by=cls.user,
		)
		OpeningBalanceImportItem.objects.create(
			opening_balance=cls.opening_balance,
			item=cls.item,
			location=cls.location,
			batch_lot="RO-BATCH-001",
			quantity=Decimal("10"),
			unit_price=Decimal("100"),
			sumber_dana=cls.funding,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=cls.item,
			location=cls.location,
			batch_lot="RO-BATCH-001",
			quantity=Decimal("10"),
			unit_price=Decimal("100"),
			sumber_dana=cls.funding,
			reference_type=Transaction.ReferenceType.INITIAL_IMPORT,
			reference_id=cls.opening_balance.pk,
			user=cls.user,
		)

	def setUp(self):
		self.client.force_login(self.user)

	def _category_row(self, response):
		return response.context["rekap_data"][0]["categories"][0]

	def test_rekap_classifies_effective_opening_import_as_saldo_awal(self):
		response = self.client.get(
			reverse("reports:rekap"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = self._category_row(response)
		self.assertEqual(row["saldo_awal"], Decimal("1000"))
		self.assertEqual(row["nilai_terima"], Decimal("0"))
		self.assertEqual(row["saldo_akhir"], Decimal("1000"))

	def test_rekap_preserves_large_precise_values_in_post_processing(self):
		precise_total = Decimal("99999999999891234567891.008765432109")
		OpeningBalanceImportItem.objects.filter(
			opening_balance=self.opening_balance,
			item=self.item,
		).update(
			quantity=Decimal("9999999999.99"),
			unit_price=Decimal("9999999999999.1234567891"),
		)
		Transaction.objects.filter(
			reference_type=Transaction.ReferenceType.INITIAL_IMPORT,
			reference_id=self.opening_balance.pk,
			item=self.item,
		).update(
			quantity=Decimal("9999999999.99"),
			unit_price=Decimal("9999999999999.1234567891"),
		)

		response = self.client.get(
			reverse("reports:rekap"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = self._category_row(response)
		group = response.context["rekap_data"][0]
		self.assertEqual(row["saldo_awal"], precise_total)
		self.assertEqual(row["saldo_akhir"], precise_total)
		self.assertEqual(group["subtotal_saldo_awal"], precise_total)
		self.assertEqual(group["subtotal_saldo_akhir"], precise_total)
		self.assertEqual(response.context["grand_totals"]["saldo_awal"], precise_total)
		self.assertEqual(response.context["grand_totals"]["saldo_akhir"], precise_total)
		self.assertContains(
			response,
			"Rp 99.999.999.999.891.234.567.891,008765432109",
		)
		self.assertNotContains(
			response,
			"Rp 99.999.999.999.891.234.567.891,01",
		)

	def test_rekap_negates_large_precise_outbound_values_without_rounding(self):
		precise_total = Decimal("99999999999891234567891.008765432109")
		expected_saldo_akhir = Decimal("-99999999999891234566891.008765432109")
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.OUT,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-OUT",
			quantity=Decimal("9999999999.99"),
			unit_price=Decimal("9999999999999.1234567891"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.DISTRIBUTION,
			reference_id=999,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:rekap"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = self._category_row(response)
		group = response.context["rekap_data"][0]
		self.assertEqual(row["nilai_distribusi"], precise_total)
		self.assertEqual(row["saldo_awal"], Decimal("1000"))
		self.assertEqual(row["saldo_akhir"], expected_saldo_akhir)
		self.assertEqual(group["subtotal_saldo_akhir"], expected_saldo_akhir)
		self.assertEqual(
			response.context["grand_totals"]["saldo_akhir"],
			expected_saldo_akhir,
		)

	def test_rekap_keeps_legacy_unlinked_initial_import_as_saldo_awal(self):
		legacy_tx = Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LEGACY",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.INITIAL_IMPORT,
			reference_id=999999,
			user=self.user,
		)
		Transaction.objects.filter(pk=legacy_tx.pk).update(
			created_at=timezone.make_aware(datetime(2026, 1, 1, 8, 0))
		)

		response = self.client.get(
			reverse("reports:rekap"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = self._category_row(response)
		self.assertEqual(row["saldo_awal"], Decimal("1700"))
		self.assertEqual(row["nilai_terima"], Decimal("0"))
		self.assertEqual(row["saldo_akhir"], Decimal("1700"))

	def test_rekap_keeps_legacy_initial_import_when_reference_id_collides_with_future_header(self):
		future_header = OpeningBalanceImport.objects.create(
			document_number="SALDO-AWAL-2027-FUTURE",
			effective_date=date(2027, 1, 1),
			created_by=self.user,
		)
		legacy_tx = Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LEGACY-COLLIDE",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.INITIAL_IMPORT,
			reference_id=future_header.pk,
			user=self.user,
		)
		Transaction.objects.filter(pk=legacy_tx.pk).update(
			created_at=timezone.make_aware(datetime(2026, 1, 1, 8, 0))
		)

		response = self.client.get(
			reverse("reports:rekap"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = self._category_row(response)
		self.assertEqual(row["saldo_awal"], Decimal("1700"))
		self.assertEqual(row["nilai_terima"], Decimal("0"))
		self.assertEqual(row["saldo_akhir"], Decimal("1700"))

	def test_rekap_keeps_in_period_legacy_initial_import_as_received(self):
		legacy_tx = Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LEGACY-IN-PERIOD",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.INITIAL_IMPORT,
			reference_id=999998,
			user=self.user,
		)
		Transaction.objects.filter(pk=legacy_tx.pk).update(
			created_at=timezone.make_aware(datetime(2026, 1, 15, 8, 0))
		)

		response = self.client.get(
			reverse("reports:rekap"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = self._category_row(response)
		self.assertEqual(row["saldo_awal"], Decimal("1000"))
		self.assertEqual(row["nilai_terima"], Decimal("700"))
		self.assertEqual(row["saldo_akhir"], Decimal("1700"))

	def test_detailed_report_keeps_in_period_legacy_initial_import_as_received(self):
		legacy_tx = Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LEGACY-DETAIL",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.INITIAL_IMPORT,
			reference_id=999997,
			user=self.user,
		)
		Transaction.objects.filter(pk=legacy_tx.pk).update(
			created_at=timezone.make_aware(datetime(2026, 1, 15, 8, 0))
		)

		response = self.client.get(
			reverse("reports:index"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = next(
			row for row in response.context["report_data"]
			if row["batch_lot"] == "RO-BATCH-LEGACY-DETAIL"
		)
		self.assertEqual(row["initial_stock"], Decimal("0"))
		self.assertEqual(row["received"], Decimal("7"))
		self.assertEqual(row["ending_stock"], Decimal("7"))

	def test_rekap_classifies_in_period_linked_opening_import_as_received(self):
		in_period_opening = OpeningBalanceImport.objects.create(
			document_number="SALDO-AWAL-2026-IN-PERIOD",
			effective_date=date(2026, 1, 15),
			created_by=self.user,
		)
		OpeningBalanceImportItem.objects.create(
			opening_balance=in_period_opening,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LINKED-IN-PERIOD",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LINKED-IN-PERIOD",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.INITIAL_IMPORT,
			reference_id=in_period_opening.pk,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:rekap"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = self._category_row(response)
		self.assertEqual(row["saldo_awal"], Decimal("1000"))
		self.assertEqual(row["nilai_terima"], Decimal("700"))
		self.assertEqual(row["saldo_akhir"], Decimal("1700"))

	def test_detailed_report_classifies_in_period_linked_opening_import_as_received(self):
		in_period_opening = OpeningBalanceImport.objects.create(
			document_number="SALDO-AWAL-2026-IN-PERIOD-DETAIL",
			effective_date=date(2026, 1, 15),
			created_by=self.user,
		)
		OpeningBalanceImportItem.objects.create(
			opening_balance=in_period_opening,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LINKED-IN-PERIOD-DETAIL",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LINKED-IN-PERIOD-DETAIL",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.INITIAL_IMPORT,
			reference_id=in_period_opening.pk,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:index"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = next(
			row for row in response.context["report_data"]
			if row["batch_lot"] == "RO-BATCH-LINKED-IN-PERIOD-DETAIL"
		)
		self.assertEqual(row["initial_stock"], Decimal("0"))
		self.assertEqual(row["received"], Decimal("7"))
		self.assertEqual(row["ending_stock"], Decimal("7"))

	def test_detailed_report_separates_same_batch_price_by_source_document(self):
		Stock.objects.create(
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-SOURCE-SPLIT",
			source_document_number="SALDO-AWAL-SOURCE-A",
			expiry_date=date(2030, 1, 1),
			quantity=Decimal("5"),
			reserved=Decimal("0"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
		)
		Stock.objects.create(
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-SOURCE-SPLIT",
			source_document_number="SALDO-AWAL-SOURCE-B",
			expiry_date=date(2031, 1, 1),
			quantity=Decimal("7"),
			reserved=Decimal("0"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-SOURCE-SPLIT",
			source_document_number="SALDO-AWAL-SOURCE-A",
			quantity=Decimal("5"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=11,
			user=self.user,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-SOURCE-SPLIT",
			source_document_number="SALDO-AWAL-SOURCE-B",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=12,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:index"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		rows = sorted(
			(
				row["source_document_number"],
				row["expiry_date"],
				row["received"],
			)
			for row in response.context["report_data"]
			if row["batch_lot"] == "RO-BATCH-SOURCE-SPLIT"
		)
		self.assertEqual(
			rows,
			[
				("SALDO-AWAL-SOURCE-A", date(2030, 1, 1), Decimal("5")),
				("SALDO-AWAL-SOURCE-B", date(2031, 1, 1), Decimal("7")),
			],
		)
		self.assertContains(response, "SALDO-AWAL-SOURCE-A")
		self.assertContains(response, "SALDO-AWAL-SOURCE-B")

		export_response = self.client.get(
			reverse("reports:index"),
			{
				"start_date": "2026-01-01",
				"end_date": "2026-12-31",
				"format": "excel",
			},
		)
		self.assertEqual(export_response.status_code, 200)
		workbook = load_workbook(BytesIO(export_response.content))
		sheet = workbook.active
		self.assertEqual(sheet["D4"].value, "Lokasi")
		self.assertEqual(sheet["F4"].value, "Dokumen Sumber")
		source_values = {
			sheet.cell(row=row_idx, column=6).value
			for row_idx in range(5, sheet.max_row + 1)
		}
		self.assertIn("SALDO-AWAL-SOURCE-A", source_values)
		self.assertIn("SALDO-AWAL-SOURCE-B", source_values)

	def test_detailed_report_displays_exact_high_precision_unit_prices(self):
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-EXACT-PRICE",
			source_document_number="RCV-EXACT-PRICE-A",
			quantity=Decimal("5"),
			unit_price=Decimal("1000.1234567890"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=21,
			user=self.user,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-EXACT-PRICE",
			source_document_number="RCV-EXACT-PRICE-B",
			quantity=Decimal("7"),
			unit_price=Decimal("1000.1240000000"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=22,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:index"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "1.000,123456789")
		self.assertContains(response, "1.000,124")
		self.assertNotContains(response, "1000.12")

	def test_detailed_report_nets_receiving_reversal_rows(self):
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-RCV-CORR",
			source_document_number="RCV-CORR-001",
			quantity=Decimal("10"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=41,
			user=self.user,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.OUT,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-RCV-CORR",
			source_document_number="RCV-CORR-001",
			quantity=Decimal("3"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=41,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:index"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = next(
			row
			for row in response.context["report_data"]
			if row["batch_lot"] == "RO-BATCH-RCV-CORR"
		)
		self.assertEqual(row["received"], Decimal("7"))
		self.assertEqual(row["ending_stock"], Decimal("7"))

	def test_detailed_report_separates_same_source_batch_by_location_expiry(self):
		other_location = Location.objects.create(
			code="RO-LOC-OTHER",
			name="Gudang Pembantu",
		)
		Stock.objects.create(
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LOCATION-SPLIT",
			source_document_number="RCV-LOCATION-SPLIT",
			expiry_date=date(2030, 1, 1),
			quantity=Decimal("5"),
			reserved=Decimal("0"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
		)
		Stock.objects.create(
			item=self.item,
			location=other_location,
			batch_lot="RO-BATCH-LOCATION-SPLIT",
			source_document_number="RCV-LOCATION-SPLIT",
			expiry_date=date(2031, 1, 1),
			quantity=Decimal("7"),
			reserved=Decimal("0"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-LOCATION-SPLIT",
			source_document_number="RCV-LOCATION-SPLIT",
			quantity=Decimal("5"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=21,
			user=self.user,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=other_location,
			batch_lot="RO-BATCH-LOCATION-SPLIT",
			source_document_number="RCV-LOCATION-SPLIT",
			quantity=Decimal("7"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=22,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:index"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		rows = sorted(
			(
				row["location_label"],
				row["expiry_date"],
				row["received"],
			)
			for row in response.context["report_data"]
			if row["batch_lot"] == "RO-BATCH-LOCATION-SPLIT"
		)
		self.assertEqual(
			rows,
			[
				("RO-LOC - Gudang Rekap", date(2030, 1, 1), Decimal("5")),
				("RO-LOC-OTHER - Gudang Pembantu", date(2031, 1, 1), Decimal("7")),
			],
		)
		self.assertContains(response, "RO-LOC-OTHER - Gudang Pembantu")

	def test_detailed_report_groups_locations_by_unique_location_id(self):
		same_name_location = Location.objects.create(
			code="RO-LOC-DUP",
			name=self.location.name,
		)
		for location, quantity in (
			(self.location, Decimal("5")),
			(same_name_location, Decimal("7")),
		):
			Stock.objects.create(
				item=self.item,
				location=location,
				batch_lot="RO-BATCH-DUP-LOCATION",
				source_document_number="RCV-DUP-LOCATION",
				expiry_date=date(2030, 1, 1),
				quantity=quantity,
				reserved=Decimal("0"),
				unit_price=Decimal("100"),
				sumber_dana=self.funding,
			)
			Transaction.objects.create(
				transaction_type=Transaction.TransactionType.IN,
				item=self.item,
				location=location,
				batch_lot="RO-BATCH-DUP-LOCATION",
				source_document_number="RCV-DUP-LOCATION",
				quantity=quantity,
				unit_price=Decimal("100"),
				sumber_dana=self.funding,
				reference_type=Transaction.ReferenceType.RECEIVING,
				reference_id=30 + location.pk,
				user=self.user,
			)

		response = self.client.get(
			reverse("reports:index"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		rows = sorted(
			(row["location_label"], row["received"])
			for row in response.context["report_data"]
			if row["batch_lot"] == "RO-BATCH-DUP-LOCATION"
		)
		self.assertEqual(
			rows,
			[
				("RO-LOC - Gudang Rekap", Decimal("5")),
				("RO-LOC-DUP - Gudang Rekap", Decimal("7")),
			],
		)
		self.assertContains(response, "RO-LOC-DUP - Gudang Rekap")

	def test_detailed_report_applies_in_period_transfers_by_location(self):
		destination = Location.objects.create(
			code="RO-LOC-TRANSFER",
			name="Gudang Transfer",
		)
		Stock.objects.create(
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-TRANSFER",
			source_document_number="RCV-TRANSFER-SOURCE",
			expiry_date=date(2030, 1, 1),
			quantity=Decimal("6"),
			reserved=Decimal("0"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
		)
		Stock.objects.create(
			item=self.item,
			location=destination,
			batch_lot="RO-BATCH-TRANSFER",
			source_document_number="RCV-TRANSFER-SOURCE",
			expiry_date=date(2030, 1, 1),
			quantity=Decimal("4"),
			reserved=Decimal("0"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-TRANSFER",
			source_document_number="RCV-TRANSFER-SOURCE",
			quantity=Decimal("10"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=31,
			user=self.user,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.OUT,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-TRANSFER",
			source_document_number="RCV-TRANSFER-SOURCE",
			quantity=Decimal("4"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.TRANSFER,
			reference_id=32,
			user=self.user,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=destination,
			batch_lot="RO-BATCH-TRANSFER",
			source_document_number="RCV-TRANSFER-SOURCE",
			quantity=Decimal("4"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.TRANSFER,
			reference_id=32,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:index"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		rows = {
			row["location_label"]: row
			for row in response.context["report_data"]
			if row["batch_lot"] == "RO-BATCH-TRANSFER"
		}
		self.assertEqual(rows["RO-LOC - Gudang Rekap"]["received"], Decimal("10"))
		self.assertEqual(rows["RO-LOC - Gudang Rekap"]["transfer_out"], Decimal("4"))
		self.assertEqual(rows["RO-LOC - Gudang Rekap"]["ending_stock"], Decimal("6"))
		self.assertEqual(rows["RO-LOC-TRANSFER - Gudang Transfer"]["transfer_in"], Decimal("4"))
		self.assertEqual(rows["RO-LOC-TRANSFER - Gudang Transfer"]["ending_stock"], Decimal("4"))
		self.assertContains(response, "Transfer In")
		self.assertContains(response, "Transfer Out")

	def test_rekap_next_year_carries_prior_year_ending_balance_without_reimport(self):
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-002",
			quantity=Decimal("5"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=1,
			user=self.user,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.OUT,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-001",
			quantity=Decimal("3"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.DISTRIBUTION,
			reference_id=1,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:rekap"),
			{"start_date": "2027-01-01", "end_date": "2027-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = self._category_row(response)
		self.assertEqual(row["saldo_awal"], Decimal("1200"))
		self.assertEqual(row["nilai_terima"], Decimal("0"))
		self.assertEqual(row["saldo_akhir"], Decimal("1200"))

	def test_rekap_nets_receiving_reversal_values(self):
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.IN,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-RCV-CORR",
			source_document_number="RCV-CORR-REKAP-001",
			quantity=Decimal("10"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=42,
			user=self.user,
		)
		Transaction.objects.create(
			transaction_type=Transaction.TransactionType.OUT,
			item=self.item,
			location=self.location,
			batch_lot="RO-BATCH-RCV-CORR",
			source_document_number="RCV-CORR-REKAP-001",
			quantity=Decimal("3"),
			unit_price=Decimal("100"),
			sumber_dana=self.funding,
			reference_type=Transaction.ReferenceType.RECEIVING,
			reference_id=42,
			user=self.user,
		)

		response = self.client.get(
			reverse("reports:rekap"),
			{"start_date": "2026-01-01", "end_date": "2026-12-31"},
		)

		self.assertEqual(response.status_code, 200)
		row = self._category_row(response)
		self.assertEqual(row["saldo_awal"], Decimal("1000"))
		self.assertEqual(row["nilai_terima"], Decimal("700"))
		self.assertEqual(row["saldo_akhir"], Decimal("1700"))


class PengeluaranReportTests(TestCase):
	@classmethod
	def setUpTestData(cls):
		cls.user = User.objects.create_superuser(
			username="pengeluaran-admin",
			password="secret12345",
		)
		cls.unit = Unit.objects.create(code="BOT", name="Botol")
		cls.category = Category.objects.create(
			code="OUT-CAT", name="Outbound Category", sort_order=1
		)
		cls.item = Item.objects.create(
			nama_barang="Amoxicillin Syrup",
			satuan=cls.unit,
			kategori=cls.category,
		)
		cls.location = Location.objects.create(code="OUT-LOC", name="Gudang Pengeluaran")
		cls.funding_source = FundingSource.objects.create(code="DAU", name="DAU")
		cls.facility = Facility.objects.create(code="PKM-OUT", name="Puskesmas Pengeluaran")
		cls.other_facility = Facility.objects.create(code="PKM-ALT", name="Puskesmas Alternatif")
		cls.stock = Stock.objects.create(
			item=cls.item,
			location=cls.location,
			batch_lot="OUT-01",
			expiry_date="2027-10-31",
			quantity=50,
			reserved=0,
			unit_price=2500,
			sumber_dana=cls.funding_source,
		)

	def setUp(self):
		self.client.force_login(self.user)

	def _create_distribution(self, distribution_type, facility=None, document_number=None):
		dist = Distribution.objects.create(
			distribution_type=distribution_type,
			document_number=document_number or "",
			request_date="2026-04-15",
			facility=facility or self.facility,
			status=Distribution.Status.DISTRIBUTED,
			created_by=self.user,
			notes="Pengeluaran terverifikasi",
		)
		dist.items.create(
			item=self.item,
			quantity_requested=7,
			quantity_approved=5,
			stock=self.stock,
		)
		return dist

	def test_pengeluaran_report_filters_by_distribution_type(self):
		allocation_dist = self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			document_number="ALLOC-REP-001",
		)
		self._create_distribution(
			Distribution.DistributionType.LPLPO,
			document_number="LPLPO-REP-001",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
				'distribution_type': Distribution.DistributionType.ALLOCATION,
			},
			secure=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, allocation_dist.document_number)
		self.assertNotContains(response, 'LPLPO-REP-001')

	def test_pengeluaran_report_combined_view_remains_available(self):
		allocation_dist = self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			document_number="ALLOC-REP-ALL",
		)
		lplpo_dist = self._create_distribution(
			Distribution.DistributionType.LPLPO,
			document_number="LPLPO-REP-ALL",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
			},
			secure=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, allocation_dist.document_number)
		self.assertContains(response, lplpo_dist.document_number)
		self.assertContains(response, 'Semua Distribusi')
		self.assertContains(response, 'Permintaan Khusus')
		self.assertContains(response, 'Alokasi')
		self.assertContains(response, 'LPLPO')

	def test_pengeluaran_report_combines_facility_and_distribution_type_filters(self):
		matching_dist = self._create_distribution(
			Distribution.DistributionType.SPECIAL_REQUEST,
			facility=self.facility,
			document_number="SPEC-REP-001",
		)
		self._create_distribution(
			Distribution.DistributionType.SPECIAL_REQUEST,
			facility=self.other_facility,
			document_number="SPEC-REP-ALT",
		)
		self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			facility=self.facility,
			document_number="ALLOC-REP-FAC",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
				'facility': self.facility.pk,
				'distribution_type': Distribution.DistributionType.SPECIAL_REQUEST,
			},
			secure=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, matching_dist.document_number)
		self.assertNotContains(response, 'SPEC-REP-ALT')
		self.assertNotContains(response, 'ALLOC-REP-FAC')

	def test_pengeluaran_report_invalid_distribution_type_keeps_report_empty(self):
		self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			document_number="ALLOC-REP-INVALID",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
				'distribution_type': 'NOT_A_REAL_TYPE',
			},
			secure=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertFalse(response.context['form'].is_valid())
		self.assertEqual(response.context['report_data'], [])
		self.assertNotContains(response, 'ALLOC-REP-INVALID')

	def test_pengeluaran_report_excel_export_uses_active_tab_label(self):
		self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			document_number="ALLOC-REP-EXPORT",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
				'distribution_type': Distribution.DistributionType.ALLOCATION,
				'format': 'excel',
			},
			secure=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(
			response['Content-Type'],
			'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		)
		self.assertIn(
			'Laporan_Pengeluaran_Alokasi_2026-04-01_2026-04-30.xlsx',
			response['Content-Disposition'],
		)

	def test_pengeluaran_report_tabs_include_distribution_type_query(self):
		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
			},
			secure=True,
		)

		self.assertEqual(response.status_code, 200)
		tabs = response.context['tabs']
		self.assertTrue(any(tab['value'] == '' for tab in tabs))
		self.assertTrue(any(tab['value'] == Distribution.DistributionType.ALLOCATION for tab in tabs))

		allocation_tab = next(
			tab for tab in tabs if tab['value'] == Distribution.DistributionType.ALLOCATION
		)
		self.assertIn('distribution_type=ALLOCATION', allocation_tab['url'])
		self.assertIn('start_date=2026-04-01', allocation_tab['url'])
		self.assertIn('end_date=2026-04-30', allocation_tab['url'])

	def test_pengeluaran_excel_neutralizes_formula_prefixed_text_and_keeps_numeric_cells(self):
		response = export_pengeluaran_excel(
			[
				{
					"document_number": "=DOC-OUT-1",
					"facility_name": "+Puskesmas Formula",
					"nama_barang": "@Amoxicillin",
					"satuan": "-Botol",
					"batch_lot": "=BATCH-01",
					"expiry_date": None,
					"sumber_dana": "=DAU",
					"unit_price": Decimal("9999999999999.1234567891"),
					"quantity": Decimal("1.01"),
					"total_price": Decimal("10099999999999.114691356991"),
				}
			],
			"2026-04-01",
			"2026-04-30",
			facility_name="=Semua Fasilitas",
			distribution_type_label="+Semua Distribusi",
		)

		workbook = load_workbook(BytesIO(response.content))
		sheet = workbook.active

		self.assertEqual(
			sheet["A2"].value,
			"Periode: 2026-04-01 s/d 2026-04-30 | Fasilitas: =Semua Fasilitas | Jenis Distribusi: +Semua Distribusi",
		)
		self.assertEqual(sheet["B5"].value, "'=DOC-OUT-1")
		self.assertEqual(sheet["C5"].value, "'+Puskesmas Formula")
		self.assertEqual(sheet["D5"].value, "'@Amoxicillin")
		self.assertEqual(sheet["E5"].value, "'-Botol")
		self.assertEqual(sheet["F5"].value, "'=BATCH-01")
		self.assertEqual(sheet["H5"].value, "'=DAU")
		self.assertEqual(sheet["A2"].data_type, "s")
		self.assertEqual(sheet["I5"].value, "9999999999999.1234567891")
		self.assertEqual(sheet["J5"].value, 1.01)
		self.assertEqual(sheet["K5"].value, "10099999999999.114691356991")
		self.assertEqual(sheet["I5"].data_type, "s")
		self.assertEqual(sheet["J5"].data_type, "n")
		self.assertEqual(sheet["K5"].data_type, "s")
		self.assertEqual(sheet["K6"].value, "10099999999999.114691356991")

	def test_rekap_excel_preserves_negative_decimal_text_without_escape_prefix(self):
		response = export_rekap_excel(
			[
				{
					"sd_name": "DAU",
					"subtotal_saldo_awal": Decimal("-1.23"),
					"subtotal_nilai_terima": Decimal("0"),
					"subtotal_nilai_distribusi": Decimal("0"),
					"subtotal_nilai_ed": Decimal("0"),
					"subtotal_saldo_akhir": Decimal("-1.23"),
					"categories": [
						{
							"kategori": "Obat",
							"saldo_awal": Decimal("-1.23"),
							"nilai_terima": Decimal("0"),
							"nilai_distribusi": Decimal("0"),
							"nilai_ed": Decimal("0"),
							"saldo_akhir": Decimal("-1.23"),
						}
					],
				}
			],
			{
				"saldo_awal": Decimal("-1.23"),
				"nilai_terima": Decimal("0"),
				"nilai_distribusi": Decimal("0"),
				"nilai_ed": Decimal("0"),
				"saldo_akhir": Decimal("-1.23"),
			},
			"2026-04-01",
			"2026-04-30",
		)

		workbook = load_workbook(BytesIO(response.content))
		sheet = workbook.active

		self.assertEqual(sheet["C5"].value, "-1.23")
		self.assertEqual(sheet["C6"].value, "-1.23")
		self.assertEqual(sheet["C7"].value, "-1.23")
		self.assertEqual(sheet["G7"].value, "-1.23")
		self.assertEqual(sheet["C5"].data_type, "s")


class ProcurementReportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_superuser(
            username="proc-report-admin",
            password="secret12345",
        )
        cls.unit = Unit.objects.create(code="PRT", name="Pcs")
        cls.category = Category.objects.create(code="PRC", name="Procurement Category", sort_order=3)
        cls.item = Item.objects.create(
            kode_barang="REP-PROC-001",
            nama_barang="Vitamin C 500mg",
            satuan=cls.unit,
            kategori=cls.category,
        )
        cls.location = Location.objects.create(code="REP-PROC-LOC", name="Gudang Pengadaan")
        cls.funding = FundingSource.objects.create(code="APBD", name="APBD")
        cls.supplier = Supplier.objects.create(code="REP-SUP", name="PT Report Supplier")
        cls.contract = ProcurementContract.objects.create(
            document_number="SPJ-2026-00077",
            contract_date="2026-07-01",
            supplier=cls.supplier,
            sumber_dana=cls.funding,
            status=ProcurementContract.Status.APPROVED,
            created_by=cls.user,
            approved_by=cls.user,
        )
        cls.receiving = Receiving.objects.create(
            document_number="RCV-2026-00999",
            receiving_type=Receiving.ReceivingType.PROCUREMENT,
            receiving_date="2026-07-10",
            supplier=cls.supplier,
            sumber_dana=cls.funding,
            status=Receiving.Status.VERIFIED,
            is_planned=False,
            contract=cls.contract,
            created_by=cls.user,
            verified_by=cls.user,
        )
        ReceivingItem.objects.create(
            receiving=cls.receiving,
            item=cls.item,
            quantity=5,
            batch_lot="REP-BATCH-001",
            expiry_date="2030-01-01",
            unit_price=1200,
            location=cls.location,
            received_by=cls.user,
        )

    def setUp(self):
        self.client.force_login(self.user)

    def test_procurement_report_shows_contract_reference(self):
        response = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31"},
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "SPJ-2026-00077")
        self.assertEqual(response.context["report_data"][0]["contract_document_number"], "SPJ-2026-00077")

    def test_procurement_report_excel_includes_spj_column(self):
        response = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31", "format": "excel"},
            secure=True,
        )

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet["C4"].value, "No. SPJ")
        self.assertEqual(sheet["C5"].value, "SPJ-2026-00077")

    def test_procurement_report_excel_calculates_large_totals_with_widened_precision(self):
        ReceivingItem.objects.filter(receiving=self.receiving, item=self.item).update(
            quantity=Decimal("9999999999.99"),
            unit_price=Decimal("9999999999999.1234567891"),
        )

        response = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31", "format": "excel"},
            secure=True,
        )

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet["I5"].value, "9999999999999.1234567891")
        self.assertEqual(sheet["K5"].value, "99999999999891234567891.008765432109")
        self.assertEqual(sheet["K6"].value, "99999999999891234567891.008765432109")

    def test_procurement_report_html_displays_exact_prices_and_values(self):
        ReceivingItem.objects.filter(receiving=self.receiving, item=self.item).update(
            quantity=Decimal("9999999999.99"),
            unit_price=Decimal("9999999999999.1234567891"),
        )

        response = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31"},
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rp 9.999.999.999.999,1234567891")
        self.assertContains(
            response,
            "Rp 99.999.999.999.891.234.567.891,008765432109",
        )
        self.assertNotContains(
            response,
            '<td class="text-end">Rp 9.999.999.999.999,12</td>',
            html=True,
        )

    def test_procurement_report_excludes_cancelled_receiving(self):
        self.receiving.status = Receiving.Status.CANCELLED
        self.receiving.save(update_fields=["status", "updated_at"])

        response = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31"},
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["report_data"], [])
        self.assertNotContains(response, self.receiving.document_number)


class ProcurementReceivingReportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_superuser(
            username="report-proc-admin",
            password="secret12345",
        )
        cls.unit = Unit.objects.create(code="PRT", name="Pcs")
        cls.category = Category.objects.create(
            code="PROC-REP",
            name="Procurement Report",
            sort_order=2,
        )
        cls.item = Item.objects.create(
            kode_barang="PROC-REP-001",
            nama_barang="Rapid Test",
            satuan=cls.unit,
            kategori=cls.category,
            minimum_stock=0,
        )
        cls.location = Location.objects.create(code="PROC-REP-LOC", name="Gudang Report")
        cls.funding_source = FundingSource.objects.create(code="PROC-REP-FS", name="Procurement FS")
        cls.supplier = Supplier.objects.create(code="PROC-REP-SUP", name="Supplier Report")

    def setUp(self):
        self.client.force_login(self.user)

    def test_pengadaan_report_includes_contract_reference(self):
        from apps.procurement.models import ProcurementContract, ProcurementContractLine
        from apps.procurement.services import approve_contract
        from apps.receiving.models import Receiving, ReceivingOrderItem

        contract = ProcurementContract.objects.create(
            document_number="",
            contract_date=date(2026, 6, 28),
            supplier=self.supplier,
            sumber_dana=self.funding_source,
            notes="Kontrak report",
            created_by=self.user,
            status=ProcurementContract.Status.SUBMITTED,
            submitted_by=self.user,
            submitted_at=timezone.now(),
        )
        line = ProcurementContractLine.objects.create(
            contract=contract,
            item=self.item,
            original_quantity=Decimal("8"),
            original_unit_price=Decimal("11000"),
        )
        approve_contract(contract, self.user)
        receiving = Receiving.objects.get(contract=contract)
        order_item = ReceivingOrderItem.objects.get(receiving=receiving, contract_line=line)
        receiving.receiving_date = date(2026, 6, 28)
        receiving.save(update_fields=["receiving_date", "updated_at"])

        response = self.client.post(
            reverse("receiving:receiving_plan_receive", args=[receiving.pk]),
            {
                "items-TOTAL_FORMS": "1",
                "items-INITIAL_FORMS": "0",
                "items-MIN_NUM_FORMS": "0",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-order_item": str(order_item.pk),
                "items-0-quantity": "8",
                "items-0-batch_lot": "PROC-REPORT-BATCH",
                "items-0-expiry_date": "2031-12-31",
                "items-0-unit_price": "11000",
                "items-0-location": str(self.location.pk),
            },
            secure=True,
        )
        self.assertEqual(response.status_code, 302)
        receipt = ReceivingItem.objects.get(receiving=receiving)
        receipt.received_at = timezone.make_aware(datetime(2026, 7, 10, 9, 0, 0))
        receipt.save(update_fields=["received_at"])

        page = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31"},
            secure=True,
        )

        self.assertEqual(page.status_code, 200)
        self.assertContains(page, contract.document_number)
        self.assertContains(page, receiving.document_number)
        self.assertContains(page, "NO. SPJ")
        self.assertEqual(page.context["report_data"][0]["receiving_date"], date(2026, 7, 10))

    def test_pengadaan_report_includes_partial_procurement_receipts(self):
        from apps.receiving.models import Receiving

        contract = ProcurementContract.objects.create(
            document_number="SPJ-2026-00123",
            contract_date=date(2026, 7, 1),
            supplier=self.supplier,
            sumber_dana=self.funding_source,
            notes="Kontrak partial report",
            created_by=self.user,
            status=ProcurementContract.Status.APPROVED,
            approved_by=self.user,
            approved_at=timezone.now(),
        )
        receiving = Receiving.objects.create(
            document_number="RCV-2026-00123",
            receiving_type=Receiving.ReceivingType.PROCUREMENT,
            receiving_date=date(2026, 7, 1),
            supplier=self.supplier,
            sumber_dana=self.funding_source,
            status=Receiving.Status.PARTIAL,
            is_planned=True,
            contract=contract,
            created_by=self.user,
            approved_by=self.user,
            approved_at=timezone.now(),
        )
        ReceivingItem.objects.create(
            receiving=receiving,
            item=self.item,
            quantity=Decimal("4"),
            batch_lot="PROC-PARTIAL-BATCH",
            expiry_date=date(2031, 12, 31),
            unit_price=Decimal("12000"),
            location=self.location,
            received_by=self.user,
            received_at=timezone.make_aware(datetime(2026, 7, 11, 10, 0, 0)),
        )

        page = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31"},
            secure=True,
        )

        self.assertEqual(page.status_code, 200)
        self.assertContains(page, contract.document_number)
        self.assertContains(page, receiving.document_number)
        self.assertTrue(
            any(
                row["contract_document_number"] == contract.document_number
                and row["receiving_date"] == date(2026, 7, 11)
                and row["quantity"] == Decimal("4")
                for row in page.context["report_data"]
            )
        )
